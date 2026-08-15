"""Excel (.xlsx) parsing — one segment per sheet, rows serialized as text."""
from __future__ import annotations

from pathlib import Path

from app.parsers.base import PageSegment, ParseError

_MAX_ROWS_PER_SHEET = 2000


def parse(path: Path) -> list[PageSegment]:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ParseError(f"Не удалось прочитать Excel-файл: {exc}") from exc

    segments: list[PageSegment] = []
    for page, sheet in enumerate(wb.worksheets, start=1):
        lines: list[str] = []
        header: list[str] | None = None
        for r, row in enumerate(sheet.iter_rows(values_only=True)):
            if r >= _MAX_ROWS_PER_SHEET:
                break
            values = [str(c) for c in row if c is not None]
            if not values:
                continue
            if header is None:
                header = values
                lines.append(" | ".join(values))
            else:
                if len(values) == len(header):
                    lines.append(", ".join(f"{h}: {v}" for h, v in zip(header, values)))
                else:
                    lines.append(" | ".join(values))
        text = "\n".join(lines).strip()
        if text:
            segments.append(PageSegment(page=page, text=text, label=sheet.title))
    wb.close()

    if not segments:
        raise ParseError("Excel-файл пуст.")
    return segments
