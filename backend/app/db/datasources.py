"""Data sources registry — decouples Oleg from the hardcoded RideGo schema.

A *data source* bundles together everything Oleg needs to answer questions:
  - a human-readable name and description
  - a schema catalog (the text injected into the LLM prompt)
  - a SQLAlchemy engine pointing at the actual database

The built-in ``ridego`` source is always available and backed by the seeded demo
DB. User-uploaded files (CSV or Excel) are stored as tables in a separate SQLite
database (``data/csv_sources.db``) and registered here so Oleg can query them
with the same machinery.

Excel specifics: each **sheet** of a ``.xlsx`` workbook becomes its own data
source (e.g. a workbook with sheets "Sales" and "Customers" yields two sources).
Excel-native types (int/float/datetime) are mapped directly to SQLite affinity
instead of being guessed from strings.

Persistence: source metadata lives in the app database (``app.db``) alongside
scenarios and feedback. The uploaded data itself lives in ``csv_sources.db``.
"""
from __future__ import annotations

import csv
import io
import re
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any

from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine

from app.config import get_settings
from app.db import app_db
from app.db.schema_catalog import SCHEMA_CATALOG as RIDEGO_CATALOG
from app.db.seed import get_engine as get_ridego_engine

# The built-in source id. Always present; cannot be deleted.
RIDEGO_SOURCE_ID = "ridego"

# SQLite column type affinity chosen per detected column kind.
_SQLITE_TYPE = {
    "integer": "INTEGER",
    "real": "REAL",
    "text": "TEXT",
}

# Cap on uploaded file size and ingested rows per source.
_MAX_ROWS = 50_000


# --------------------------------------------------------------------------- #
# Public API — metadata is persisted in app.db (see app_db module)
# --------------------------------------------------------------------------- #


def list_sources() -> list[dict[str, Any]]:
    """Return metadata for all known sources (for the UI selector)."""
    out = []
    for s in app_db.list_datasources():
        out.append(
            {
                "id": s["id"],
                "name": s["name"],
                "kind": s["kind"],
                "description": s.get("description", ""),
                "row_count": s.get("row_count"),
                "created_at": s.get("created_at"),
            }
        )
    return out


def get_source_meta(source_id: str) -> dict[str, Any] | None:
    return app_db.get_datasource(source_id)


def get_engine_for(source_id: str) -> Engine:
    """Return the SQLAlchemy engine backing ``source_id``."""
    meta = get_source_meta(source_id)
    if meta is None:
        raise KeyError(f"Unknown data source: {source_id!r}")
    if meta["kind"] == "ridego":
        return get_ridego_engine()
    if meta["kind"] == "csv":
        return _uploaded_engine()
    if meta["kind"] == "postgres":
        return _postgres_engine(meta)
    raise ValueError(f"Unsupported source kind: {meta['kind']!r}")


def get_schema_catalog(source_id: str) -> str:
    """Return the schema text to inject into Oleg's prompt for ``source_id``."""
    meta = get_source_meta(source_id)
    if meta is None:
        raise KeyError(f"Unknown data source: {source_id!r}")
    if meta["kind"] == "ridego":
        return RIDEGO_CATALOG
    if meta["kind"] == "csv":
        return _build_uploaded_catalog(meta)
    if meta["kind"] == "postgres":
        # Catalog is cached in metadata at registration time.
        return _build_postgres_catalog(meta)
    raise ValueError(f"Unsupported source kind: {meta['kind']!r}")


def get_dialect(source_id: str) -> str:
    """Return 'sqlite' or 'postgresql' for the source (for dialect-aware prompts)."""
    meta = get_source_meta(source_id)
    if meta is None:
        return "sqlite"
    return "postgresql" if meta.get("kind") == "postgres" else "sqlite"


def delete_source(source_id: str) -> None:
    """Delete a user source. The built-in ``ridego`` source cannot be removed."""
    if source_id == RIDEGO_SOURCE_ID:
        raise ValueError("The built-in RideGo source cannot be deleted.")
    meta = get_source_meta(source_id)
    if meta is None:
        raise KeyError(f"Unknown data source: {source_id!r}")
    if meta.get("table_name") and meta.get("kind") == "csv":
        # Best-effort table drop; ignore failures (table may already be gone).
        try:
            eng = _uploaded_engine()
            with eng.begin() as conn:
                conn.execute(text(f'DROP TABLE IF EXISTS "{meta["table_name"]}"'))
        except Exception:  # noqa: BLE001
            pass
    # postgres sources have no local tables to drop — just remove metadata.
    app_db.delete_datasource_row(source_id)


# --------------------------------------------------------------------------- #
# Uploaded-data SQLite engine (shared by CSV and Excel sources)
# --------------------------------------------------------------------------- #


def _uploaded_db_path() -> Path:
    return get_settings().data_dir / "csv_sources.db"


def _uploaded_engine() -> Engine:
    return create_engine(f"sqlite:///{_uploaded_db_path()}")


# --------------------------------------------------------------------------- #
# Ingestion core — format-agnostic
# --------------------------------------------------------------------------- #


def ingest_rows(
    display_name: str,
    header: list[str],
    rows: list[list[Any]],
    *,
    description: str | None = None,
) -> dict[str, Any]:
    """Create a SQLite table from already-parsed ``header`` and ``rows`` and
    register a data source. This is the shared core used by both CSV and Excel
    adapters. Returns the new source's metadata.

    ``rows`` may contain typed values (int/float/datetime/str from Excel) or
    strings (from CSV); :func:`_infer_col_kind` handles both.
    """
    if not display_name:
        raise ValueError("Display name is required.")
    header = [h.strip() for h in header]
    if not header or any(not h for h in header):
        raise ValueError("Header contains empty column names.")
    header = [_sanitize_col(h, i) for i, h in enumerate(header)]

    rows = rows[:_MAX_ROWS]
    # Normalise row widths to match the header.
    normalised: list[list[Any]] = []
    for r in rows:
        if len(r) < len(header):
            r = list(r) + [None] * (len(header) - len(r))
        elif len(r) > len(header):
            r = list(r)[: len(header)]
        normalised.append(r)

    if not normalised:
        raise ValueError("No data rows found.")

    col_types = [
        _infer_col_kind([row[i] for row in normalised]) for i in range(len(header))
    ]
    sql_types = [_SQLITE_TYPE[t] for t in col_types]

    source_id = f"csv_{uuid.uuid4().hex[:10]}"
    table_name = f"src_{source_id}"

    eng = _uploaded_engine()
    with eng.begin() as conn:
        cols_sql = ", ".join(f'"{h}" {sql_types[i]}' for i, h in enumerate(header))
        conn.execute(text(f'CREATE TABLE "{table_name}" ({cols_sql})'))
        placeholders = ", ".join(f":c{i}" for i in range(len(header)))
        insert_sql = f'INSERT INTO "{table_name}" VALUES ({placeholders})'
        batch = [_make_param_dict(header, col_types, row) for row in normalised]
        conn.execute(text(insert_sql), batch)

    meta_entry = {
        "id": source_id,
        "name": display_name,
        "kind": "csv",
        "description": description or f"Загружен источник: {len(header)} колонок, {len(normalised)} строк.",
        "table_name": table_name,
        "columns": [
            {"name": header[i], "type": col_types[i], "sqlite_type": sql_types[i]}
            for i in range(len(header))
        ],
        "row_count": len(normalised),
        "created_at": datetime.utcnow().isoformat(timespec="seconds"),
    }
    app_db.save_datasource(meta_entry)
    return meta_entry


# --------------------------------------------------------------------------- #
# CSV adapter
# --------------------------------------------------------------------------- #


def ingest_csv(
    raw_name: str,
    rows_text: str,
    *,
    delimiter: str = ",",
) -> dict[str, Any]:
    """Parse CSV text and ingest it as a single data source."""
    if not raw_name:
        raise ValueError("File name is required.")
    reader = csv.reader(rows_text.splitlines(), delimiter=delimiter)
    try:
        header = next(reader)
    except StopIteration as e:
        raise ValueError("CSV file is empty (no header row).") from e

    data_rows: list[list[str]] = list(reader)
    return ingest_rows(_human_name(raw_name), header, data_rows)


# --------------------------------------------------------------------------- #
# Excel (.xlsx) adapter — each sheet becomes its own data source
# --------------------------------------------------------------------------- #


def ingest_xlsx(raw_name: str, file_bytes: bytes) -> list[dict[str, Any]]:
    """Parse an ``.xlsx`` workbook and ingest every non-empty sheet as its own
    data source. Returns a list of metadata dicts (one per sheet ingested).

    Raises ``ValueError`` if the file is unreadable or contains no data on any
    sheet.
    """
    from openpyxl import load_workbook

    if not raw_name:
        raise ValueError("File name is required.")

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:  # openpyxl raises a variety of error types
        raise ValueError(f"Could not read .xlsx file: {e}") from e

    file_stem = _human_name(raw_name)
    created: list[dict[str, Any]] = []
    try:
        for sheet in wb.worksheets:
            header, rows = _read_sheet(sheet)
            if header is None or not rows:
                # Skip empty sheets (no header or header-only).
                continue
            display_name = f"{file_stem} · {sheet.title}" if sheet.title else file_stem
            meta = ingest_rows(
                display_name,
                header,
                rows,
                description=f"Лист «{sheet.title}» из {raw_name}: {len(rows)} строк.",
            )
            created.append(meta)
    finally:
        wb.close()

    if not created:
        raise ValueError("The workbook contains no data on any sheet.")
    return created


def _read_sheet(sheet) -> tuple[list[str] | None, list[list[Any]]]:
    """Read a worksheet as (header, rows). Returns (None, []) for empty sheets.

    Uses ``iter_rows(values_only=True)`` so cells arrive as native Python types
    (int/float/datetime/str). Datetimes are normalised to ISO strings so the
    schema-catalog and SQL round-trip cleanly.
    """
    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        return None, []
    header = [("" if v is None else str(v)) for v in all_rows[0]]
    if not any(header):
        return None, []

    data: list[list[Any]] = []
    for raw_row in all_rows[1:]:
        if raw_row is None:
            continue
        # A row is "empty" only if every cell is None/blank.
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in raw_row):
            continue
        data.append([_normalise_cell(v) for v in raw_row])
    return header, data


def _normalise_cell(v: Any) -> Any:
    """Normalise an openpyxl cell value for SQLite storage."""
    if v is None:
        return None
    if isinstance(v, bool):
        # bool is a subclass of int in Python — keep it as text to avoid surprises.
        return str(v)
    if isinstance(v, datetime):
        # openpyxl returns date-only cells as datetime with a 00:00:00 time part;
        # store those as plain ISO dates to avoid the noisy T00:00:00 suffix.
        if v.hour == 0 and v.minute == 0 and v.second == 0 and v.microsecond == 0:
            return v.date().isoformat()
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return v


# --------------------------------------------------------------------------- #
# Helpers — naming, column sanitising, type inference
# --------------------------------------------------------------------------- #

_INT_RE = re.compile(r"^-?\d+$")
_FLOAT_RE = re.compile(r"^-?\d+[.,]\d+$")
_ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _sanitize_col(name: str, idx: int) -> str:
    cleaned = re.sub(r"[^A-Za-zА-Яа-яЁё0-9_ ]", "_", name).strip().replace(" ", "_")
    if not cleaned:
        cleaned = f"col_{idx}"
    if cleaned[0].isdigit():
        cleaned = f"c_{cleaned}"
    return cleaned.lower()


def _human_name(raw_name: str) -> str:
    base = Path(raw_name).stem
    return base.replace("_", " ").replace("-", " ").title() or raw_name


def _infer_col_kind(values: list[Any]) -> str:
    """Infer a column kind from its sample.

    Values may be typed (int/float/str from Excel) or strings (from CSV). Typed
    values are trusted directly; strings fall back to regex heuristics. Empty
    values (None/"") are ignored when deciding; if ALL values are empty the
    column defaults to text.
    """
    non_empty = [v for v in values if v is not None and v != ""]
    if not non_empty:
        return "text"

    # If every non-empty value is already a Python number, trust the runtime type.
    if all(isinstance(v, int) and not isinstance(v, bool) for v in non_empty):
        return "integer"
    if all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in non_empty):
        return "real"

    # Otherwise (strings, or mixed) — apply string heuristics on stringified values.
    str_values = [str(v) for v in non_empty]
    if all(_ISO_DATE_RE.match(v) for v in str_values):
        return "text"  # keep dates as text to preserve formatting
    if all(_INT_RE.match(v) for v in str_values):
        return "integer"
    if all(_INT_RE.match(v) or _FLOAT_RE.match(v) for v in str_values):
        return "real"
    return "text"


def _coerce_value(kind: str, raw: Any) -> Any:
    """Coerce a raw cell value into the target column kind for SQLite."""
    if raw is None or raw == "":
        return None
    # If the value is already typed (from Excel), keep numbers as-is.
    if isinstance(raw, bool):
        return str(raw)
    if kind == "integer":
        if isinstance(raw, int):
            return raw
        if isinstance(raw, float):
            return int(raw) if raw.is_integer() else raw
        try:
            return int(str(raw))
        except ValueError:
            return raw
    if kind == "real":
        if isinstance(raw, (int, float)):
            return float(raw)
        try:
            return float(str(raw).replace(",", "."))
        except ValueError:
            return raw
    return raw


def _make_param_dict(header: list[str], col_types: list[str], row: list[Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for i in range(len(header)):
        out[f"c{i}"] = _coerce_value(col_types[i], row[i] if i < len(row) else None)
    return out


def _build_uploaded_catalog(meta: dict[str, Any]) -> str:
    """Compose a schema-catalog text for an uploaded source, describing its
    single table and columns with inferred types. This is what Oleg sees."""
    table = meta["table_name"]
    cols = meta.get("columns") or []
    lines = [f"# Uploaded data source: {meta['name']}", ""]
    lines.append(f"## {table}")
    lines.append(f"Rows: {meta.get('row_count', '?')}. Column types below.")
    for c in cols:
        kind = c.get("type", "text")
        lines.append(f"- {c['name']} {kind.upper()}")
    lines.append("")
    lines.append("Notes:")
    lines.append("- Use SQLite syntax. Quote table/column names if unsure (they are case-sensitive).")
    lines.append("- Date columns are stored as TEXT in ISO format (YYYY-MM-DD).")
    lines.append("- Only SELECT / WITH queries are allowed.")
    return "\n".join(lines)


# --------------------------------------------------------------------------- #
# PostgreSQL sources
# --------------------------------------------------------------------------- #


def _build_postgres_url(conn: dict[str, Any]) -> str:
    """Build a SQLAlchemy URL from connection parameters."""
    return (
        f"postgresql+psycopg2://{conn['username']}:{conn['password']}"
        f"@{conn['host']}:{conn.get('port', 5432)}/{conn['database']}"
    )


def _postgres_engine(meta: dict[str, Any]) -> Engine:
    """Create a SQLAlchemy engine for a postgres source."""
    conn = meta.get("connection") or {}
    if not conn.get("host"):
        raise ValueError("Postgres source missing connection info.")
    return create_engine(_build_postgres_url(conn), pool_pre_ping=True)


def _introspect_schema(engine: Engine, max_tables: int = 50) -> list[dict[str, Any]]:
    """Introspect table/column metadata using the dialect-agnostic SQLAlchemy
    ``inspect()`` API. Works identically for SQLite and PostgreSQL."""
    from sqlalchemy import inspect as sa_inspect

    insp = sa_inspect(engine)
    tables: list[dict[str, Any]] = []
    for table_name in insp.get_table_names()[:max_tables]:
        cols = []
        for col in insp.get_columns(table_name):
            cols.append({
                "name": col["name"],
                "type": str(col["type"]),
                "nullable": col.get("nullable", True),
            })
        tables.append({"name": table_name, "columns": cols})
    return tables


def _build_postgres_catalog(meta: dict[str, Any]) -> str:
    """Build a schema-catalog text from cached introspection results."""
    tables = meta.get("columns") or []
    conn = meta.get("connection") or {}
    lines = [f"# PostgreSQL source: {meta['name']}", ""]
    lines.append(f"Database: {conn.get('database', '?')} @ {conn.get('host', '?')}")
    lines.append("")
    for t in tables:
        lines.append(f"## {t['name']}")
        for c in t["columns"]:
            lines.append(f"- {c['name']} {c['type']}")
        lines.append("")
    lines.append("Notes:")
    lines.append("- Use PostgreSQL syntax (DATE_TRUNC, EXTRACT, ::cast, ILIKE).")
    lines.append("- Only SELECT / WITH queries are allowed.")
    return "\n".join(lines)


def register_postgres(
    *,
    name: str,
    host: str,
    port: int = 5432,
    database: str,
    username: str,
    password: str,
    source_id: str | None = None,
) -> dict[str, Any]:
    """Test the connection, introspect the schema, and register a postgres source.

    Raises ``ValueError`` if the connection fails or psycopg2 is unavailable.
    """
    try:
        engine = create_engine(
            f"postgresql+psycopg2://{username}:{password}@{host}:{port}/{database}",
            pool_pre_ping=True,
        )
        # Test the connection + introspect in one shot.
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        tables = _introspect_schema(engine)
    except Exception as e:
        raise ValueError(f"Could not connect to PostgreSQL: {e}") from e

    sid = source_id or f"pg_{uuid.uuid4().hex[:10]}"
    conn_info = {"host": host, "port": port, "database": database, "username": username, "password": password}
    meta = {
        "id": sid,
        "name": name,
        "kind": "postgres",
        "description": f"PostgreSQL: {database}@{host}, {len(tables)} таблиц(ы).",
        "table_name": None,
        "columns": tables,  # cached introspection results
        "row_count": None,
        "created_at": datetime.utcnow().isoformat(timespec="seconds"),
        "connection": conn_info,
    }
    app_db.save_datasource(meta)
    # Return a copy with masked password for the API response.
    safe = dict(meta)
    safe["connection"] = {**conn_info, "password": "****"}
    return safe


def refresh_postgres_schema(source_id: str) -> dict[str, Any]:
    """Re-introspect the schema for an existing postgres source."""
    meta = get_source_meta(source_id)
    if meta is None:
        raise KeyError(f"Unknown data source: {source_id!r}")
    if meta["kind"] != "postgres":
        raise ValueError("Source is not a PostgreSQL source.")
    engine = _postgres_engine(meta)
    tables = _introspect_schema(engine)
    meta["columns"] = tables
    meta["description"] = f"PostgreSQL: {meta['connection']['database']}@{meta['connection']['host']}, {len(tables)} таблиц(ы)."
    app_db.save_datasource(meta)
    safe = dict(meta)
    safe["connection"] = {**meta["connection"], "password": "****"}
    return safe


def register_env_postgres() -> dict[str, Any] | None:
    """If POSTGRES_URL is set, register/update a postgres source on startup."""
    url = get_settings().postgres_url
    if not url:
        return None
    # Parse the URL: postgresql://user:pass@host:port/db
    import re
    from urllib.parse import urlparse

    parsed = urlparse(url)
    if not parsed.hostname:
        return None
    try:
        return register_postgres(
            name=f"PostgreSQL ({parsed.path.lstrip('/') or parsed.hostname})",
            host=parsed.hostname,
            port=parsed.port or 5432,
            database=parsed.path.lstrip("/"),
            username=parsed.username or "",
            password=parsed.password or "",
            source_id="postgres-env",
        )
    except ValueError:
        # Connection failed at startup — skip silently (user can add manually later).
        return None
